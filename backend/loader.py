"""
Excel-based data loader for ABET Dashboard.
Reads .xlsx files from backend/data/ and computes student outcome percentages.

Directory structure expected:
  data/
    Cycle 1/
      Program Name/
        2024B/   ← .xlsx files with datetime-encoded SO headers
        2025A/   ← .xlsx files with datetime-encoded SO headers
        (or) ABET 2024B/, ABET 2025A/
    Cycle 2/
      Program Name/
        ← .xlsx files directly (all 2026A), headers like "1.1"
"""

import logging
import re
from collections import defaultdict
from pathlib import Path
from typing import Optional

import openpyxl

from models import (
    CourseRow,
    MetricSet,
    PeriodOutcomes,
    ALL_SUB_OUTCOME_CODES,
    parse_percentage,
)

logger = logging.getLogger(__name__)

# ── Course name normalisation ───────────────────────────────────────

# Courses that appear under different names across semesters
COURSE_ALIASES: dict[str, str] = {
    "infraestructure and computer security": "Infrastructure and Computer Security",
    "infrastructure and computer security": "Infrastructure and Computer Security",
    "technological negotiation models": "Technological Negotiation Models",
    "algorithmics and object oriented programming ii": "Algorithmics and Object Oriented Programming II",
    "telematic systems": "Telematic Systems",
    "it management": "IT Management",
    "gestión de ti": "IT Management",
    "software engineering i": "Software Engineering I",
    "software engineering ii": "Software Engineering II",
    "software engineering": "Software Engineering",
    "cloud computing": "Cloud Computing",
    "information engineering": "Information Engineering",
    "objects and data structures": "Objects and Data Structures",
    # Civil Eng. aliases
    "construction planning and scheduling": "Construction Planning and Scheduling",
    "construction planning and scheduling ": "Construction Planning and Scheduling",
    "geometric design of roads": "Geometric Design of Roads",
    "geometric design of road": "Geometric Design of Roads",
    "materials lab": "Materials Lab",
    "mechanics of materials": "Mechanics of Materials",
    "mecánica de materiales": "Mechanics of Materials",
    "design of concrete structures": "Design of Concrete Structures",
    "design of concrete structures i": "Design of Concrete Structures I",
    "diseño de estructuras de concreto ii": "Design of Concrete Structures II",
    "soils and construction materials lab": "Soils and Construction Materials Lab",
    "soils and conctruction materials lab": "Soils and Construction Materials Lab",
    "geographic information systems": "Geographic Information Systems",
    # Industrial Eng. aliases
    "complex thinking": "Complex Thinking",
    "modeling and simulation": "Modeling and Simulation",
    "experimental methods": "Experimental Methods",
    "operations management i": "Operations Management I",
    "operations management ii": "Operations Management II",
    "operations research i": "Operations Research I",
    "operation research i": "Operations Research I",
    "human factor engineering": "Human Factor Engineering",
    "quality engineering": "Quality Engineering",
    "project planning and evaluation": "Project Planning and Evaluation",
    "supply chain design": "Supply Chain Design",
}


def normalise_course_name(raw: str) -> str:
    """Normalise a course name for cross-semester matching."""
    cleaned = raw.strip().lower()
    return COURSE_ALIASES.get(cleaned, raw.strip())


def extract_path_info(file_path: str, data_dir: str = "data") -> tuple[str, str, str]:
    """
    Extract (cycle, program, semester) from a file path relative to data_dir.
    
    Handles variations:
      - Cycle 1/Civil Eng./2024B/file.xlsx  → ("Cycle 1", "Civil Eng.", "2024B")
      - Cycle 1/Systems Eng./ABET 2024B/..  → ("Cycle 1", "Systems Eng.", "2024B")
      - Cycle 2/Civil Eng./file.xlsx        → ("Cycle 2", "Civil Eng.", "2026A")
    """
    path = Path(file_path)
    base = Path(data_dir)
    try:
        rel = path.relative_to(base)
    except ValueError:
        rel = path
    parts = rel.parts

    # First part is the cycle folder
    cycle = parts[0] if len(parts) > 0 else "Unknown"
    # Second part is the program name
    program = parts[1] if len(parts) > 1 else "Unknown"

    # parts structure:
    #   4+ parts: Cycle X / Program / semester_folder / filename.xlsx
    #   3 parts:  Cycle X / Program / filename.xlsx  (no semester subfolder → Cycle 2 = 2026A)
    if len(parts) >= 4:
        # There is a semester subfolder → could be "2024B", "2025A", "ABET 2024B", etc.
        semester_raw = parts[2]
        # Extract semester pattern like 2024B, 2025A, 2026A from names like "ABET 2024B"
        match = re.search(r'(\d{4}[AB])', semester_raw, re.IGNORECASE)
        if match:
            semester = match.group(1)
        else:
            semester = semester_raw
    else:
        # Files directly under program folder → Cycle 2 = 2026A
        if "Cycle 2" in cycle or "cycle 2" in cycle.lower():
            semester = "2026A"
        else:
            semester = "Unknown"

    return cycle, program, semester


def normalise_professor(raw: str) -> str:
    """Clean up professor name, preserving original casing."""
    return raw.strip() if raw else ""


# ── Low-level Excel parsing ─────────────────────────────────────────

def decode_so_header(cell_value) -> Optional[str]:
    """
    Decode a header cell into an SO sub-outcome code like "1.1".
    
    Cycle 1: datetime(year, month, day) → SO code = f"{day}.{month}"
    Cycle 2: numeric like 1.1, 2.3, etc.
    """
    if cell_value is None:
        return None
    if hasattr(cell_value, 'month') and hasattr(cell_value, 'day'):
        # Cycle 1: datetime encoding
        so_num = int(cell_value.day)
        sub_num = int(cell_value.month)
        return f"{so_num}.{sub_num}"
    # Cycle 2: direct numeric or string
    s = str(cell_value).strip()
    # try to match a pattern like "1.1", "2.3", etc.
    parts = s.replace(",", ".").split(".")
    if len(parts) == 2:
        try:
            int(parts[0])
            int(parts[1])
            return s
        except ValueError:
            pass
    return None


def _read_header_row(ws, row_num: int = 1) -> list:
    """Read a header row, skipping leading empty cells."""
    raw_headers = []
    started = False
    for cell in ws[row_num]:
        if cell.value is None:
            if started:
                break
            else:
                raw_headers.append(None)
                continue
        started = True
        raw_headers.append(cell.value)
    return raw_headers


def _has_any_value(row_data: list) -> bool:
    """Check if a row has any non-None value."""
    return any(v is not None for v in row_data)


def parse_cycle1_file(path: str) -> dict:
    """
    Parse a Cycle 1 (2024B/2025A) Excel file.
    
    Returns a dict with:
      - course: normalised course name
      - professor: cleaned professor name
      - semester: e.g. "2024B"
      - scores: { student_id: { so_code: score } }
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Parse header row (try row 1 first, fall back to row 2 if empty)
    raw_headers = _read_header_row(ws, 1)
    header_row_num = 1
    if not _has_any_value(raw_headers):
        raw_headers = _read_header_row(ws, 2)
        header_row_num = 2

    # Identify columns
    so_columns: list[tuple[int, str]] = []  # (col_index, so_code)
    col_course: Optional[int] = None
    col_prof: Optional[int] = None
    col_student: Optional[int] = None

    # Collect column candidates (prefer exact matches over partial)
    student_candidates: list[tuple[int, int]] = []   # (col_index, priority: 0=exact, 1=partial)
    course_candidates: list[tuple[int, int]] = []
    prof_candidates: list[tuple[int, int]] = []

    for i, h in enumerate(raw_headers):
        so_code = decode_so_header(h)
        if so_code:
            so_columns.append((i, so_code))
        elif isinstance(h, str):
            hl = h.strip().lower()
            # Student ID column: prefer exact "cod estudiante" over just "estudiante"
            if hl in ('cod estudiante', 'codigo estudiante'):
                student_candidates.append((i, 0))
            elif 'estudiante' in hl or 'student' in hl:
                student_candidates.append((i, 1))
            # Course column: prefer exact "asignatura" over "cod asignatura"
            if hl in ('asignatura', 'course'):
                course_candidates.append((i, 0))
            elif 'asignatura' in hl or 'course' in hl:
                course_candidates.append((i, 1))
            # Professor column
            if hl in ('profesor', 'professor'):
                prof_candidates.append((i, 0))
            elif 'profesor' in hl or 'professor' in hl:
                prof_candidates.append((i, 1))

    # Pick best candidate (lowest priority = most exact match, then first occurrence)
    student_candidates.sort(key=lambda x: (x[1], x[0]))
    course_candidates.sort(key=lambda x: (x[1], x[0]))
    prof_candidates.sort(key=lambda x: (x[1], x[0]))
    col_student = student_candidates[0][0] if student_candidates else None
    col_course = course_candidates[0][0] if course_candidates else None
    col_prof = prof_candidates[0][0] if prof_candidates else None

    # Try to parse student_id from various formats
    SKIP_WORDS = {'average', 'promedio', 'total', 'nota', 'grade'}
    def _parse_student_id(val) -> Optional[str]:
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return str(int(val))
        s = str(val).strip()
        if not s:
            return None
        # Skip summary rows (but NOT group labels like "Grupo1")
        if s.lower() in SKIP_WORDS:
            return None
        # Try numeric conversion first
        try:
            return str(int(float(s)))
        except (ValueError, TypeError):
            pass
        # Try extracting digits (e.g. "25A33" → "2533")
        digits = re.sub(r'\D', '', s)
        if digits and len(digits) >= 3:
            return digits
        # Fallback: use original if it looks like an ID (e.g. "Grupo1", "Grupo2")
        return s if len(s) >= 3 else None

    # Collect data
    course_name = None
    professor = None
    scores: dict[str, dict[str, float]] = defaultdict(dict)

    for row_idx in range(header_row_num + 1, ws.max_row + 1):
        row_vals = [ws.cell(row=row_idx, column=c + 1).value for c in range(len(raw_headers))]

        student_id = _parse_student_id(row_vals[col_student]) if col_student is not None else None
        if student_id is None:
            continue

        # Extract course name (may be empty in some rows → inherit from previous)
        if col_course is not None and row_vals[col_course]:
            course_name = normalise_course_name(str(row_vals[col_course]))

        # Extract professor
        if col_prof is not None and row_vals[col_prof]:
            professor = normalise_professor(str(row_vals[col_prof]))

        # Extract SO scores
        for ci, so_code in so_columns:
            val = row_vals[ci]
            if val is not None:
                # Skip text values that are clearly not scores (long descriptions, etc.)
                if isinstance(val, str) and len(val.strip()) > 10:
                    continue
                try:
                    score = float(val)
                    if 1 <= score <= 5:
                        scores[student_id][so_code] = score
                except (ValueError, TypeError):
                    pass

    # If course name not found in data, derive from filename
    if not course_name:
        stem = Path(path).stem
        # Remove trailing spaces
        course_name = normalise_course_name(stem)

    # Derive cycle, program, and semester from path
    _, program, semester = extract_path_info(path)

    return {
        "course": course_name,
        "program": program,
        "professor": professor or "",
        "semester": semester,
        "scores": dict(scores),
    }


def parse_cycle2_file(path: str) -> dict:
    """
    Parse a Cycle 2 (2026A) Excel file.
    Direct numeric headers like "1.1", "2.3", etc.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Parse header row (try row 1 first, fall back to row 2 if empty)
    raw_headers = _read_header_row(ws, 1)
    header_row_num = 1
    if not _has_any_value(raw_headers):
        raw_headers = _read_header_row(ws, 2)
        header_row_num = 2

    so_columns: list[tuple[int, str]] = []
    col_course: Optional[int] = None
    col_prof: Optional[int] = None
    col_student: Optional[int] = None

    # Collect column candidates (prefer exact matches over partial)
    student_candidates: list[tuple[int, int]] = []
    course_candidates: list[tuple[int, int]] = []
    prof_candidates: list[tuple[int, int]] = []

    for i, h in enumerate(raw_headers):
        so_code = decode_so_header(h)
        if so_code:
            so_columns.append((i, so_code))
        elif isinstance(h, str):
            hl = h.strip().lower()
            if hl in ('cod estudiante', 'codigo estudiante'):
                student_candidates.append((i, 0))
            elif 'estudiante' in hl or 'student' in hl:
                student_candidates.append((i, 1))
            if hl in ('asignatura', 'course'):
                course_candidates.append((i, 0))
            elif 'asignatura' in hl or 'course' in hl:
                course_candidates.append((i, 1))
            if hl in ('profesor', 'professor'):
                prof_candidates.append((i, 0))
            elif 'profesor' in hl or 'professor' in hl:
                prof_candidates.append((i, 1))

    student_candidates.sort(key=lambda x: (x[1], x[0]))
    course_candidates.sort(key=lambda x: (x[1], x[0]))
    prof_candidates.sort(key=lambda x: (x[1], x[0]))
    col_student = student_candidates[0][0] if student_candidates else None
    col_course = course_candidates[0][0] if course_candidates else None
    col_prof = prof_candidates[0][0] if prof_candidates else None

    # Robust student ID parser
    SKIP_WORDS = {'average', 'promedio', 'total', 'nota', 'grade'}
    def _parse_student_id(val) -> Optional[str]:
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return str(int(val))
        s = str(val).strip()
        if not s:
            return None
        if s.lower() in SKIP_WORDS:
            return None
        try:
            return str(int(float(s)))
        except (ValueError, TypeError):
            pass
        digits = re.sub(r'\D', '', s)
        if digits and len(digits) >= 3:
            return digits
        return s if len(s) >= 3 else None

    course_name = None
    professor = None
    scores: dict[str, dict[str, float]] = defaultdict(dict)

    for row_idx in range(header_row_num + 1, ws.max_row + 1):
            course_name = normalise_course_name(str(row_vals[col_course]))

        if col_prof is not None and row_vals[col_prof]:
            professor = normalise_professor(str(row_vals[col_prof]))

        for ci, so_code in so_columns:
            val = row_vals[ci]
            if val is not None:
                # Skip text values that are clearly not scores (long descriptions, headers, etc.)
                if isinstance(val, str) and len(val.strip()) > 10:
                    continue
                try:
                    score = float(val)
                    if 1 <= score <= 5:
                        scores[student_id][so_code] = score
                except (ValueError, TypeError):
                    pass

    if not course_name:
        stem = Path(path).stem.rstrip()
        course_name = normalise_course_name(stem)

    # Derive cycle, program, and semester from path
    _, program, semester = extract_path_info(path)

    return {
        "course": course_name,
        "program": program,
        "professor": professor or "",
        "semester": semester,
        "scores": dict(scores),
    }


def parse_file(path: str) -> dict:
    """Parse an Excel file, auto-detecting Cycle 1 vs Cycle 2 format."""
    # Cycle 2 files have direct numeric headers; Cycle 1 have datetime headers
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    first_header = ws.cell(row=1, column=2).value  # skip student ID col

    if hasattr(first_header, 'month') and hasattr(first_header, 'day'):
        return parse_cycle1_file(path)
    else:
        return parse_cycle2_file(path)


# ── Percentage computation ──────────────────────────────────────────

def compute_percentages(scores: dict[str, dict[str, float]]) -> dict[str, MetricSet]:
    """
    Given per-student scores for a course+semester, compute % >=3, % >=4, % =5
    for each SO sub-outcome.
    """
    # Collect all scores per SO code
    so_score_lists: dict[str, list[float]] = defaultdict(list)
    for student_id, so_scores in scores.items():
        for so_code, score in so_scores.items():
            so_score_lists[so_code].append(score)

    result: dict[str, MetricSet] = {}
    for so_code, score_list in so_score_lists.items():
        n = len(score_list)
        if n == 0:
            result[so_code] = MetricSet(ge3=None, ge4=None, eq5=None)
            continue
        ge3 = round(sum(1 for s in score_list if s >= 3) / n * 100, 1)
        ge4 = round(sum(1 for s in score_list if s >= 4) / n * 100, 1)
        eq5 = round(sum(1 for s in score_list if s == 5) / n * 100, 1)
        result[so_code] = MetricSet(ge3=ge3, ge4=ge4, eq5=eq5)

    return result


# ── Main loader ─────────────────────────────────────────────────────

def load_all_courses(data_dir: str = "data") -> list[CourseRow]:
    """
    Walk the data/ directory, parse all Excel files, compute percentages,
    and merge courses across semesters into dashboard rows.
    """
    base = Path(data_dir)
    if not base.exists():
        logger.error("Data directory not found: %s", base)
        return []

    # ── Step 1: parse all files ──────────────────────────────────
    # course_name → semester → { so_code: MetricSet, professor, scores }
    course_data: dict[str, dict[str, dict]] = defaultdict(dict)

    for xlsx_path in sorted(base.rglob("*.xlsx")):
        # Skip temp files
        if xlsx_path.name.startswith("~$"):
            continue
        rel = str(xlsx_path.relative_to(base))
        logger.info("Parsing %s ...", rel)
        try:
            parsed = parse_file(str(xlsx_path))
        except Exception as e:
            logger.warning("  Failed to parse %s: %s", rel, e)
            continue

        cname = parsed["course"]
        sem = parsed["semester"]
        prof = parsed["professor"]
        scores = parsed["scores"]

        if not scores:
            logger.warning("  No scores found in %s", rel)
            continue

        pcts = compute_percentages(scores)

        if sem not in course_data[cname]:
            course_data[cname][sem] = {
                "professor": prof,
                "outcomes": {},
                "student_count": len(scores),
            }

        existing_sem = course_data[cname][sem]
        existing_sem["outcomes"].update(pcts)
        # Keep first professor found (not last)
        if prof and not existing_sem["professor"]:
            existing_sem["professor"] = prof

        logger.info("  → %s [%s]: %d students, %d SOs evaluated",
                     cname, sem, len(scores), len(pcts))

    # ── Step 2: build CourseRow objects ──────────────────────────
    # Cycle mapping
    SEMESTER_CYCLE = {
        "2024B": "Cycle 1",
        "2025A": "Cycle 1",
        "2026A": "Cycle 2",
    }

    # Collect programs per course (same course name may appear in multiple programs)
    course_programs: dict[str, set[str]] = defaultdict(set)
    for xlsx_path in sorted(base.rglob("*.xlsx")):
        if xlsx_path.name.startswith("~$"):
            continue
        _, prog, _ = extract_path_info(str(xlsx_path))
        try:
            parsed = parse_file(str(xlsx_path))
        except Exception:
            continue
        cname = parsed["course"]
        if cname:
            course_programs[cname].add(prog)

    rows: list[CourseRow] = []
    for cname in sorted(course_data.keys()):
        semesters = course_data[cname]
        # Use the first program found; join multiple with " / "
        programs = course_programs.get(cname, {"Unknown"})
        program_str = " / ".join(sorted(programs))
        row = CourseRow(
            course=cname,
            program=program_str,
            professor_2024b=semesters.get("2024B", {}).get("professor") or None,
            professor_2025a=semesters.get("2025A", {}).get("professor") or None,
            professor_2026a=semesters.get("2026A", {}).get("professor") or None,
        )

        # Store per-period data
        for sem in ["2024B", "2025A", "2026A"]:
            if sem in semesters:
                sem_data = semesters[sem]
                row.periods[sem] = PeriodOutcomes(
                    semester=sem,
                    cycle=SEMESTER_CYCLE.get(sem, "Unknown"),
                    professor=sem_data.get("professor") or None,
                    student_count=sem_data.get("student_count", 0),
                    outcomes={**sem_data["outcomes"]},
                )

        # Merge outcomes (aggregated): most recent semester data takes priority.
        for sem in ["2026A", "2025A", "2024B"]:
            if sem in semesters:
                for so_code, ms in semesters[sem]["outcomes"].items():
                    if so_code not in row.outcomes:
                        row.outcomes[so_code] = ms

        rows.append(row)

    logger.info("Total dashboard rows: %d courses", len(rows))
    return rows
